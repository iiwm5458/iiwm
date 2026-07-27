from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Iterable


HEADERS = [
    "\u5bf9\u5c40\u8f6e\u6b21",
    "\u529f\u65b9\u9009\u624b",
    "\u529f\u65b9\u9009\u624bID",
    "\u5b88\u65b9\u9009\u624b",
    "\u5b88\u65b9\u9009\u624bID",
    "\u529f\u65b9P1",
    "\u529f\u65b9P1\u6218\u529b",
    "\u529f\u65b9P2",
    "\u529f\u65b9P2\u6218\u529b",
    "\u529f\u65b9P3",
    "\u529f\u65b9P3\u6218\u529b",
    "\u529f\u65b9P4",
    "\u529f\u65b9P4\u6218\u529b",
    "\u529f\u65b9P5",
    "\u529f\u65b9P5\u6218\u529b",
    "\u5b88\u65b9P1",
    "\u5b88\u65b9P1\u6218\u529b",
    "\u5b88\u65b9P2",
    "\u5b88\u65b9P2\u6218\u529b",
    "\u5b88\u65b9P3",
    "\u5b88\u65b9P3\u6218\u529b",
    "\u5b88\u65b9P4",
    "\u5b88\u65b9P4\u6218\u529b",
    "\u5b88\u65b9P5",
    "\u5b88\u65b9P5\u6218\u529b",
    "\u80dc\u65b9",
    "\u7f6e\u4fe1\u5ea6",
    "\u6e90\u56fe\u7247",
]

ROSTER_HEADERS = [
    "\u53c2\u8d5b\u9009\u624b",
    "\u9009\u624bID",
    "\u9635\u5bb91",
    "\u9635\u5bb91\u6218\u529b",
    "\u9635\u5bb91\u6536\u85cf",
    "\u9635\u5bb92",
    "\u9635\u5bb92\u6218\u529b",
    "\u9635\u5bb92\u6536\u85cf",
    "\u9635\u5bb93",
    "\u9635\u5bb93\u6218\u529b",
    "\u9635\u5bb93\u6536\u85cf",
    "\u9635\u5bb94",
    "\u9635\u5bb94\u6218\u529b",
    "\u9635\u5bb94\u6536\u85cf",
    "\u9635\u5bb95",
    "\u9635\u5bb95\u6218\u529b",
    "\u9635\u5bb95\u6536\u85cf",
]

ROSTER_STAT_HEADERS = [
    "\u6781\u4e50\u51c0\u571f",
    "\u6cf0\u7279\u62c9",
    "\u7c73\u897f\u5229\u65af",
    "\u671d\u5723\u8005",
    "\u53cd\u5e38",
    "\u706b\u529b\u578b",
    "\u9632\u5fa1\u578b",
    "\u8f85\u52a9\u578b",
]

GROUP_STAGE_LABELS = {
    "\u0036\u0034\u8fdb\u0033\u0032": "\u0038\u8fdb\u0034",
    "\u0033\u0032\u8fdb\u0031\u0036": "\u0034\u8fdb\u0032",
    "\u0031\u0036\u8fdb\u0038": "\u0032\u8fdb\u0031",
}

CHINESE_NUMBERS = {
    "\u96f6": 0,
    "\u4e00": 1,
    "\u4e8c": 2,
    "\u4e09": 3,
    "\u56db": 4,
    "\u4e94": 5,
    "\u516d": 6,
    "\u4e03": 7,
    "\u516b": 8,
    "\u4e5d": 9,
    "\u5341": 10,
}


def _as_int(value, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        text = str(value or "").strip()
        if text in CHINESE_NUMBERS:
            return CHINESE_NUMBERS[text]
        if text.startswith("\u5341") and len(text) == 2:
            return 10 + CHINESE_NUMBERS.get(text[1], 0)
        if text.endswith("\u5341") and len(text) == 2:
            return CHINESE_NUMBERS.get(text[0], 0) * 10
        if "\u5341" in text and len(text) == 3:
            return CHINESE_NUMBERS.get(text[0], 0) * 10 + CHINESE_NUMBERS.get(text[2], 0)
        return default


def _round_suffix(stage: str, match_index, round_index) -> str:
    round_number = _as_int(round_index)
    if not round_number:
        return ""
    if stage.endswith("\u0032\u8fdb\u0031") or stage == "\u51a0\u519b\u4e89\u9738\u8d5b\u51a0\u4e9a\u519b":
        return f"G{round_number}"
    match_number = _as_int(match_index)
    if not match_number:
        return f"G{round_number}"
    return f"M{match_number}G{round_number}"


def build_record_key(record: dict) -> str:
    stage = str(record.get("stage") or "")
    match_index = record.get("match_index", "")
    round_index = record.get("round_index", "")

    if stage.startswith("\u51a0\u519b\u4e89\u9738\u8d5b"):
        if stage == "\u51a0\u519b\u4e89\u9738\u8d5b\u4e00\u56fe\u6d41":
            top8_match_index = _as_int(match_index)
            if top8_match_index == 1:
                stage = "\u51a0\u519b\u4e89\u9738\u8d5b\u51a0\u4e9a\u519b"
                match_index = 1
            elif 2 <= top8_match_index <= 3:
                stage = "\u51a0\u519b\u4e89\u9738\u8d5b\u0034\u8fdb\u0032"
                match_index = top8_match_index - 1
            elif top8_match_index >= 4:
                stage = "\u51a0\u519b\u4e89\u9738\u8d5b\u0038\u8fdb\u0034"
                match_index = top8_match_index - 3
        suffix = _round_suffix(stage, match_index, round_index)
        return f"{stage} {suffix}".strip()

    group_stage = GROUP_STAGE_LABELS.get(stage)
    if group_stage:
        group_number = _as_int(record.get("group_index", ""))
        group_part = f"\u7b2c{group_number}\u7ec4" if group_number else ""
        prefix = f"\u664b\u7ea7\u8d5b{group_part}{group_stage}"
        suffix = _round_suffix(prefix, match_index, round_index)
        return f"{prefix} {suffix}".strip()

    group_index = record.get("group_index", "")
    group_number = _as_int(group_index)
    group_part = f"\u7b2c{group_number}\u7ec4" if group_number else ""
    suffix = _round_suffix(stage, match_index, round_index)
    return f"{stage}{group_part} {suffix}".strip()


def build_headers(include_power: bool = True) -> list[str]:
    headers = HEADERS[:5]
    for side in ("\u529f\u65b9", "\u5b88\u65b9"):
        for index in range(1, 6):
            headers.append(f"{side}P{index}")
            if include_power:
                headers.append(f"{side}P{index}\u6218\u529b")
    headers.extend(HEADERS[-3:])
    return headers


def build_roster_headers(
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
) -> list[str]:
    headers = ["\u53c2\u8d5b\u9009\u624b", "\u9009\u624bID"]
    for index in range(1, 6):
        headers.append(f"\u9635\u5bb9{index}")
        if include_power:
            headers.append(f"\u9635\u5bb9{index}\u6218\u529b")
        if include_collection:
            headers.append(f"\u9635\u5bb9{index}\u6536\u85cf")
    if include_stat_levels:
        headers.extend(ROSTER_STAT_HEADERS)
    return headers


ROSTER_EXPORT_HEADERS = build_roster_headers()


def _lineup(
    names: list,
    powers: list,
    collections: list | None = None,
    include_power: bool = True,
    include_collection: bool = False,
) -> list[dict]:
    names = list(names or [])[:5] + [""] * 5
    powers = list(powers or [])[:5] + [None] * 5
    collections = list(collections or [])[:5] + [""] * 5
    lineup = []
    for index in range(5):
        item = {"\u4f4d\u7f6e": f"P{index + 1}", "\u59ae\u59ec": names[index]}
        if include_power:
            item["\u6218\u529b"] = powers[index]
        if include_collection:
            item["\u6536\u85cf"] = collections[index]
        lineup.append(item)
    return lineup


def _roster_json_entry(
    entry: dict,
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
) -> dict:
    teams = entry.get("teams") or {}
    powers = entry.get("powers") or {}
    collections = entry.get("collections") or {}
    item = {
        "\u53c2\u8d5b\u9009\u624b": entry.get("nickname", ""),
        "\u9009\u624bID": str(entry.get("player_id", "") or ""),
    }
    for index in range(1, 6):
        team = list(teams.get(index) or teams.get(str(index)) or [])[:5]
        power = list(powers.get(index) or powers.get(str(index)) or [])[:5]
        collection = list(collections.get(index) or collections.get(str(index)) or [])[:5]
        item[f"\u9635\u5bb9{index}"] = team
        if include_power:
            item[f"\u9635\u5bb9{index}\u6218\u529b"] = power
        if include_collection:
            item[f"\u9635\u5bb9{index}\u6536\u85cf"] = collection
    if include_stat_levels:
        stat_levels = list(entry.get("stat_levels") or [])[: len(ROSTER_STAT_HEADERS)]
        stat_levels += [None] * (len(ROSTER_STAT_HEADERS) - len(stat_levels))
        item.update(dict(zip(ROSTER_STAT_HEADERS, stat_levels)))
    return item


def _export_roster_json(
    roster,
    output_dir: Path,
    stem: str,
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
) -> Path | None:
    entries = _roster_entries(roster)
    if not entries:
        return None
    path = output_dir / f"{stem}_roster.json"
    payload = [
        _roster_json_entry(
            entry,
            include_power=include_power,
            include_collection=include_collection,
            include_stat_levels=include_stat_levels,
        )
        for entry in entries
    ]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def export_json(
    records: list[dict],
    output_dir: Path,
    stem: str,
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
    roster=None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{stem}_result.json"
    payload = []
    for record in records:
        item = {
            "\u5bf9\u5c40\u8f6e\u6b21": build_record_key(record),
            "\u529f\u65b9\u9009\u624b": record.get("attacker_player_nickname", ""),
            "\u529f\u65b9\u9009\u624bID": record.get("attacker_player_id", ""),
            "\u5b88\u65b9\u9009\u624b": record.get("defender_player_nickname", ""),
            "\u5b88\u65b9\u9009\u624bID": record.get("defender_player_id", ""),
            "\u529f\u65b9\u9635\u5bb9": _lineup(
                record.get("attacker_team"),
                record.get("attacker_power"),
                record.get("attacker_collection"),
                include_power=include_power,
                include_collection=include_collection,
            ),
            "\u5b88\u65b9\u9635\u5bb9": _lineup(
                record.get("defender_team"),
                record.get("defender_power"),
                record.get("defender_collection"),
                include_power=include_power,
                include_collection=include_collection,
            ),
            "\u80dc\u65b9": record.get("winner", ""),
            "\u7f6e\u4fe1\u5ea6": record.get("confidence", ""),
            "\u6e90\u56fe\u7247": record.get("source_image", ""),
        }
        if include_stat_levels:
            item["\u529f\u65b9\u5faa\u73af\u7b49\u7ea7"] = record.get("attacker_stat_levels", [])
            item["\u5b88\u65b9\u5faa\u73af\u7b49\u7ea7"] = record.get("defender_stat_levels", [])
        payload.append(item)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    _export_roster_json(
        roster,
        output_dir,
        stem,
        include_power=include_power,
        include_collection=include_collection,
        include_stat_levels=include_stat_levels,
    )
    return path


def _row(record: dict, include_power: bool = True) -> list:
    attacker = list(record.get("attacker_team") or [])[:5]
    defender = list(record.get("defender_team") or [])[:5]
    attacker_power = list(record.get("attacker_power") or [])[:5]
    defender_power = list(record.get("defender_power") or [])[:5]
    attacker += [""] * (5 - len(attacker))
    defender += [""] * (5 - len(defender))
    attacker_power += [None] * (5 - len(attacker_power))
    defender_power += [None] * (5 - len(defender_power))
    attacker_columns = []
    defender_columns = []
    for index in range(5):
        attacker_columns.append(attacker[index])
        defender_columns.append(defender[index])
        if include_power:
            attacker_columns.append(attacker_power[index])
            defender_columns.append(defender_power[index])
    return [
        build_record_key(record),
        record.get("attacker_player_nickname", ""),
        record.get("attacker_player_id", ""),
        record.get("defender_player_nickname", ""),
        record.get("defender_player_id", ""),
        *attacker_columns,
        *defender_columns,
        record.get("winner", ""),
        record.get("confidence", ""),
        record.get("source_image", ""),
    ]


def _join_roster_values(values: list) -> str:
    return " / ".join(str(value) for value in (values or []) if value not in ("", None))


def _roster_entries(roster) -> list[dict]:
    if not roster:
        return []
    if isinstance(roster, dict):
        entries = list(roster.values())
    else:
        entries = list(roster)
    return sorted(
        entries,
        key=lambda entry: (
            str(entry.get("group_index") or "").zfill(2),
            str(entry.get("match_index") or "").zfill(2),
            str(entry.get("side") or ""),
            str(entry.get("player_id") or ""),
            str(entry.get("nickname") or ""),
        ),
    )


def _roster_row(
    entry: dict,
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
) -> list:
    teams = entry.get("teams") or {}
    powers = entry.get("powers") or {}
    collections = entry.get("collections") or {}
    row = [entry.get("nickname", ""), str(entry.get("player_id", "") or "")]
    for index in range(1, 6):
        team = teams.get(index) or teams.get(str(index)) or []
        power = powers.get(index) or powers.get(str(index)) or []
        collection = collections.get(index) or collections.get(str(index)) or []
        row.append(_join_roster_values(team))
        if include_power:
            row.append(_join_roster_values(power))
        if include_collection:
            row.append(_join_roster_values(collection))
    if include_stat_levels:
        stat_levels = list(entry.get("stat_levels") or [])[: len(ROSTER_STAT_HEADERS)]
        stat_levels += [None] * (len(ROSTER_STAT_HEADERS) - len(stat_levels))
        row.extend(stat_levels)
    return row


def _format_header_row(ws, header_fill, header_font, alignment_cls) -> None:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_cls(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34


def export_excel(
    records: list[dict],
    output_dir: Path,
    stem: str,
    roster=None,
    include_power: bool = True,
    include_collection: bool = True,
    include_stat_levels: bool = True,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return export_csv(records, output_dir, stem, include_power=include_power)

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{stem}.xlsx"
    headers = build_headers(include_power=include_power)
    roster_headers = build_roster_headers(
        include_power=include_power,
        include_collection=include_collection,
        include_stat_levels=include_stat_levels,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "ArenaData"
    ws.append(headers)
    for record in records:
        ws.append(_row(record, include_power=include_power))
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            if cell.column == 4:
                continue
            cell.value = str(cell.value or "")
            cell.number_format = "@"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    _format_header_row(ws, header_fill, header_font, Alignment)
    if include_power:
        for column, header in enumerate(headers, 1):
            if not header.endswith("\u6218\u529b"):
                continue
            for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2, max_row=ws.max_row):
                cell[0].number_format = "#,##0"
                cell[0].alignment = Alignment(horizontal="right")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx in range(1, len(headers) + 1):
        letter = get_column_letter(idx)
        header = headers[idx - 1]
        if idx == 1:
            width = 28
        elif header in ("\u529f\u65b9\u9009\u624bID", "\u5b88\u65b9\u9009\u624bID"):
            width = 16
        elif header in ("\u529f\u65b9\u9009\u624b", "\u5b88\u65b9\u9009\u624b"):
            width = 20
        elif (header.startswith("\u529f\u65b9P") or header.startswith("\u5b88\u65b9P")) and not header.endswith("\u6218\u529b"):
            width = 20
        elif header.endswith("\u6218\u529b"):
            width = 18
        elif header == "\u6e90\u56fe\u7247":
            width = 32
        else:
            width = 13
        ws.column_dimensions[letter].width = width
    roster_rows = _roster_entries(roster)
    if roster_rows:
        roster_ws = wb.create_sheet("\u53c2\u8d5b\u9635\u5bb9")
        roster_ws.append(roster_headers)
        for entry in roster_rows:
            roster_ws.append(
                _roster_row(
                    entry,
                    include_power=include_power,
                    include_collection=include_collection,
                    include_stat_levels=include_stat_levels,
                )
            )
        _format_header_row(roster_ws, header_fill, header_font, Alignment)
        roster_ws.freeze_panes = "A2"
        roster_ws.auto_filter.ref = roster_ws.dimensions
        for row in roster_ws.iter_rows(min_row=2, min_col=2, max_col=2):
            row[0].value = str(row[0].value or "")
            row[0].number_format = "@"
        for idx in range(1, len(roster_headers) + 1):
            letter = get_column_letter(idx)
            header = roster_headers[idx - 1]
            if idx == 1:
                width = 22
            elif idx == 2:
                width = 16
            elif header in ROSTER_STAT_HEADERS:
                width = 12
            elif re.match(r"^\u9635\u5bb9\d+$", header):
                width = 40
            elif header.endswith("\u6218\u529b"):
                width = 28
            else:
                width = 18
            roster_ws.column_dimensions[letter].width = width
        if include_stat_levels:
            stat_start = len(roster_headers) - len(ROSTER_STAT_HEADERS) + 1
            for row in roster_ws.iter_rows(
                min_row=2,
                min_col=stat_start,
                max_col=len(roster_headers),
            ):
                for cell in row:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="top")
        for row in roster_ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column <= len(roster_headers) - (len(ROSTER_STAT_HEADERS) if include_stat_levels else 0):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)
    return path


def export_csv(records: Iterable[dict], output_dir: Path, stem: str, include_power: bool = True) -> Path:
    import csv

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{stem}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(build_headers(include_power=include_power))
        for record in records:
            writer.writerow(_row(record, include_power=include_power))
    return path
