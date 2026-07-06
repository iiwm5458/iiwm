from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook
from PIL import Image, ImageDraw

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from recognizer.image_splitter import split_input_image, split_match_block
from recognizer.result_parser import (
    ATTACKER_CARD_SLOT_CENTERS,
    COLLECTION_NONE,
    COLLECTION_PRECISE_GROUP64_FHD,
    COLLECTION_PRECISE_GROUP64_WIDE,
    COLLECTION_PRECISE_GROUP64_WIDE_BLOCK_HEIGHT_MIN,
    COLLECTION_ROW_ICON_X_OFFSETS,
    DEFENDER_CARD_SLOT_CENTERS,
    _classify_collection_icon_by_color,
    _crop_rel,
)


LABELS = ("R", "R15", "SR", "SR15", "SSR", "SSR3")
NONE_LABEL = "none"
DISPLAY_NONE = COLLECTION_NONE
NORMAL_SIZE = (48, 48)
ROW_START = 0.275
ROW_END = 0.925
DEFAULT_COLLECTION_X_HALF = 0.023
DEFAULT_COLLECTION_Y_CENTER = 0.350
DEFAULT_COLLECTION_Y_HALF = 0.073
SEARCH_DELTAS = ((0.0, 0.0),)
DIRECT_SCORE_THRESHOLD = 0.15
DIRECT_FAMILY_THRESHOLD = 0.11
DIRECT_SR_DARK_THRESHOLD = 0.26
DIRECT_SSR_DARK_THRESHOLD = 0.24
DIRECT_R_DARK_THRESHOLD = 0.30
DIRECT_ORANGE_OVERRIDE = 0.07
DIRECT_CYAN_MARGIN = 0.06
DIRECT_ORANGE_CYAN_MAX = 0.08
DIRECT_SR_WHITE_GUARD = 0.30
DIRECT_SR_DARK_WHITE_RATIO_GUARD = 1.0
DIRECT_SR_PURPLE_GUARD = 0.15
DIRECT_SR_ACTIVE_GUARD = 0.15
DIRECT_SR_LOW_DARK_GUARD = 0.285
DIRECT_SR_BRIGHT_WHITE_GUARD = 0.40
DIRECT_SR_LOW_DARK_WHITE_RATIO_GUARD = 0.66
DIRECT_SR_WEAK_PURPLE_GUARD = 0.21
DIRECT_SR_SCORE_DELTA_GUARD = -0.04
DIRECT_R_TO_SR_MARGIN = 0.015
DIRECT_R_TO_SR_DARK_MAX = 0.03
DIRECT_NONE_VETO_MARGIN = -0.18
DIRECT_R_BRIGHT_NONE_MARGIN = -0.04
DIRECT_R_BRIGHT_ACTIVE_MAX = 0.20
DIRECT_R_BRIGHT_WHITE_MIN = 0.60
DIRECT_R15_CYAN_OVERRIDE_DARK_MIN = 0.24
DIRECT_R15_CYAN_OVERRIDE_DARK_MAX = 0.34
DIRECT_R15_CYAN_OVERRIDE_WHITE_MAX = 0.46
DIRECT_R15_CYAN_OVERRIDE_MARGIN = 0.04
DIRECT_R15_CYAN_OVERRIDE_ACTIVE_MIN = 0.18
DIRECT_R15_CYAN_OVERRIDE_SCORE_MARGIN = -0.08
_GENERIC_POSITIVE_MASK_CACHE: dict[int, np.ndarray] = {}


@dataclass
class TemplateBank:
    label: str
    features: np.ndarray
    masks: np.ndarray
    denominators: np.ndarray
    count: int


@dataclass
class SlotPrediction:
    dataset: str
    player_id: str
    side: str
    group_index: int
    match_index: int
    team_index: int
    slot_index: int
    expected: str
    pure_label: str
    hybrid_label: str
    color_label: str
    best_score: float
    second_score: float
    none_score: float
    dark_ratio: float
    crop: Image.Image
    debug: dict


@dataclass
class CollectionGeometry:
    x_half: float
    y_center: float
    y_half: float
    offsets: dict[str, tuple[float, ...]]
    detailed_positions: dict[str, dict[int, dict[int, "SlotGeometry"]]]


@dataclass
class SlotGeometry:
    x_center: float
    y_center: float
    x_half: float
    y_half: float


def project_root() -> Path:
    return Path(__file__).resolve().parents[5]


def tool_root() -> Path:
    return Path(__file__).resolve().parents[1]


def normalize_player_id(value) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D+", "", text)
    return digits.zfill(8) if digits else text


def normalize_label(value) -> str:
    text = str(value or "").strip().upper()
    if not text or text in {"NONE", "NAN", COLLECTION_NONE.upper(), DISPLAY_NONE.upper()}:
        return DISPLAY_NONE
    if text == "SSR15":
        return "SSR3"
    return text if text in LABELS else DISPLAY_NONE


def label_for_manifest(value: str) -> str:
    return NONE_LABEL if normalize_label(value) == DISPLAY_NONE else normalize_label(value)


def find_boss_workbook(root: Path) -> Path:
    downloads = Path.home() / "Downloads"
    candidates = sorted(downloads.glob("cjjcBase20260627225803*.xlsx"))
    if not candidates:
        raise FileNotFoundError("Boss workbook not found in Downloads.")
    return candidates[0]


def read_boss_collections(path: Path) -> dict[tuple[str, int, int], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[1]
    labels: dict[tuple[str, int, int], str] = {}
    team_counter: dict[str, int] = defaultdict(int)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        player_id = normalize_player_id(row[0])
        if not player_id:
            continue
        team_counter[player_id] += 1
        team_index = team_counter[player_id]
        for slot_index in range(1, 6):
            labels[(player_id, team_index, slot_index)] = normalize_label(row[23 + slot_index])
    workbook.close()
    return labels


def read_block_players(excel_path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    rows = list(workbook.worksheets[0].iter_rows(min_row=2, values_only=True))
    workbook.close()
    players: list[tuple[str, str]] = []
    for row_index in range(0, min(len(rows), 160), 5):
        row = rows[row_index]
        players.append((normalize_player_id(row[2]), normalize_player_id(row[4])))
    return players


def load_collection_geometry(template_dir: Path) -> CollectionGeometry:
    offsets = {
        side: tuple(values)
        for side, values in COLLECTION_ROW_ICON_X_OFFSETS.items()
    }
    x_half = DEFAULT_COLLECTION_X_HALF
    y_center = DEFAULT_COLLECTION_Y_CENTER
    y_half = DEFAULT_COLLECTION_Y_HALF
    detailed_positions: dict[str, dict[int, dict[int, SlotGeometry]]] = {}

    manifest_path = template_dir / "manifest.json"
    calibration_path = template_dir / "position_calibration.json"
    detailed_calibration_path = template_dir / "position_calibration_detailed.json"
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        configured = manifest.get("position_calibration")
        if configured:
            calibration_path = template_dir / configured
        configured_detailed = manifest.get("position_calibration_detailed")
        if configured_detailed:
            detailed_calibration_path = template_dir / configured_detailed

    if calibration_path.exists():
        calibration = json.loads(calibration_path.read_text(encoding="utf-8"))
        recommended = calibration.get("recommended_initial_crop") or {}
        crop_y0 = recommended.get("classification_crop_y0")
        crop_y1 = recommended.get("classification_crop_y1")
        if isinstance(crop_y0, (int, float)) and isinstance(crop_y1, (int, float)) and crop_y1 > crop_y0:
            y_center = (float(crop_y0) + float(crop_y1)) / 2.0
        configured_x_half = recommended.get("classification_x_half")
        if isinstance(configured_x_half, (int, float)):
            x_half = float(configured_x_half)
        configured_offsets = recommended.get("center_offsets") or {}
        for side in ("attacker", "defender"):
            values = configured_offsets.get(side)
            if isinstance(values, list) and values:
                offsets[side] = tuple(float(value) for value in values)

    if detailed_calibration_path.exists():
        detailed = json.loads(detailed_calibration_path.read_text(encoding="utf-8"))
        for side, teams in (detailed.get("recommended") or {}).items():
            detailed_positions.setdefault(side, {})
            for team, slots in teams.items():
                team_index = int(team)
                detailed_positions[side].setdefault(team_index, {})
                for slot, values in slots.items():
                    slot_index = int(slot)
                    detailed_positions[side][team_index][slot_index] = SlotGeometry(
                        x_center=float(values["x_center"]),
                        y_center=float(values["y_center"]),
                        x_half=float(values["x_half"]),
                        y_half=float(values["y_half"]),
                    )

    return CollectionGeometry(
        x_half=x_half,
        y_center=y_center,
        y_half=y_half,
        offsets=offsets,
        detailed_positions=detailed_positions,
    )


def find_64_image(folder: Path) -> Path:
    candidates = [path for path in folder.glob("*.png") if "64" in path.name and "32" in path.name]
    if not candidates:
        raise FileNotFoundError(f"No 64-to-32 PNG found in {folder}")
    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def find_wide_folder(root: Path) -> Path:
    screenshots = root / "files-mentioned-by-the-user-qq" / "outputs" / "screenshots"
    expected_name = f"2026{chr(24180)}7{chr(26376)}1{chr(26085)}2"
    expected = screenshots / expected_name
    if expected.exists():
        return expected
    matches = [path for path in screenshots.iterdir() if path.is_dir() and path.name.endswith("1日2")]
    if not matches:
        raise FileNotFoundError("Wide screenshot folder not found.")
    return matches[0]


def find_fhd_folder(root: Path) -> Path:
    screenshots = root / "files-mentioned-by-the-user-qq" / "outputs" / "screenshots"
    matches = [path for path in screenshots.iterdir() if path.is_dir() and "1080" in path.name]
    if not matches:
        raise FileNotFoundError("1920x1080 screenshot folder not found.")
    return sorted(matches, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def default_datasets(root: Path) -> list[dict]:
    screenshots = root / "files-mentioned-by-the-user-qq" / "outputs" / "screenshots"
    wide_folder = find_wide_folder(root)
    fhd_folder = find_fhd_folder(root)
    wide_excel = screenshots / "2026-07-02" / "arena_season_images_20260702_142227_result.xlsx"
    if not wide_excel.exists():
        wide_excel = sorted((screenshots / "2026-07-02").glob("arena_season_images_*_result.xlsx"))[-1]
    return [
        {
            "name": "wide_3440",
            "image": find_64_image(wide_folder),
            "excel": wide_excel,
        },
        {
            "name": "fhd_1080",
            "image": find_64_image(fhd_folder),
            "excel": sorted(fhd_folder.glob("arena_season_images_*_result.xlsx"))[-1],
        },
    ]


def load_template(path: Path) -> tuple[np.ndarray, np.ndarray]:
    image = Image.open(path).convert("RGBA").resize(NORMAL_SIZE, Image.Resampling.LANCZOS)
    rgba = np.asarray(image, dtype=np.uint8)
    rgb = rgba[:, :, :3]
    alpha = rgba[:, :, 3] > 24
    if float(alpha.mean()) < 0.04:
        alpha[:, :] = True
    return rgb, alpha


def make_edge(rgb: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    return cv2.Canny(gray, 70, 170).astype(np.float32) / 255.0


def feature_channels(rgb: np.ndarray) -> np.ndarray:
    hsv = cv2.cvtColor(rgb.astype(np.uint8), cv2.COLOR_RGB2HSV).astype(np.float32)
    hue = hsv[:, :, 0] / 179.0 * 2.0 * np.pi
    saturation = hsv[:, :, 1] / 255.0
    value = hsv[:, :, 2] / 255.0
    rgb_f = rgb.astype(np.float32) / 255.0
    return np.stack(
        [
            np.cos(hue),
            np.sin(hue),
            saturation,
            value,
            rgb_f[:, :, 0],
            rgb_f[:, :, 1],
            rgb_f[:, :, 2],
        ],
        axis=2,
    ).astype(np.float32)


def collection_visual_stats(rgb: np.ndarray, mask: np.ndarray | None = None) -> dict[str, float]:
    rgb_i = rgb.astype(np.int16)
    r = rgb_i[:, :, 0]
    g = rgb_i[:, :, 1]
    b = rgb_i[:, :, 2]
    mx = np.maximum.reduce([r, g, b])
    mn = np.minimum.reduce([r, g, b])
    hsv = cv2.cvtColor(rgb.astype(np.uint8), cv2.COLOR_RGB2HSV)
    hue = hsv[:, :, 0]
    saturation = hsv[:, :, 1] / 255.0
    value = hsv[:, :, 2] / 255.0
    if mask is None:
        mask = np.ones(rgb.shape[:2], dtype=bool)
    active = mask & (saturation > 0.22) & (value > 0.25)
    cyan = active & (hue >= 78) & (hue <= 105)
    purple = active & (hue >= 122) & (hue <= 162)
    orange = active & (((hue >= 2) & (hue <= 28)) | ((hue >= 170) & (hue <= 179)))
    dark = mask & (value < 0.45)
    white = mask & (saturation < 0.22) & (value > 0.68)
    total = max(1, int(np.sum(mask)))
    return {
        "cyan": float(np.sum(cyan) / total),
        "purple": float(np.sum(purple) / total),
        "orange": float(np.sum(orange) / total),
        "active": float(np.sum(active) / total),
        "family_max": float(max(np.sum(cyan), np.sum(purple), np.sum(orange)) / total),
        "family_sum": float((np.sum(cyan) + np.sum(purple) + np.sum(orange)) / total),
        "dark": float(np.sum(dark) / total),
        "white": float(np.sum(white) / total),
        "dark_white_ratio": float(np.sum(dark) / max(1, np.sum(white))),
    }


def pick_evenly(items: list[dict], limit: int) -> list[dict]:
    if limit <= 0 or len(items) <= limit:
        return items
    return [items[round(index * (len(items) - 1) / max(1, limit - 1))] for index in range(limit)]


def load_manifest_templates(template_dir: Path, skip_strong: bool, per_label_limit: int) -> dict[str, list[tuple[np.ndarray, np.ndarray]]]:
    manifest = json.loads((template_dir / "manifest.json").read_text(encoding="utf-8"))
    entries_by_label: dict[str, list[dict]] = defaultdict(list)
    for entry in manifest.get("templates", []):
        label = normalize_label(entry.get("label"))
        if label == DISPLAY_NONE:
            label = NONE_LABEL
        if label not in (*LABELS, NONE_LABEL):
            continue
        if skip_strong and label != NONE_LABEL and entry.get("quality_review_priority") == "strong_suspect":
            continue
        entries_by_label[label].append(entry)
    grouped: dict[str, list[tuple[np.ndarray, np.ndarray]]] = defaultdict(list)
    for label, entries in entries_by_label.items():
        for entry in pick_evenly(entries, per_label_limit):
            relative_path = entry.get("path") or entry.get("tight_path") or entry.get("full_path")
            if not relative_path:
                continue
            path = template_dir / relative_path
            if path.exists():
                grouped[label].append(load_template(path))
    return grouped


def make_template_banks(grouped: dict[str, list[tuple[np.ndarray, np.ndarray]]]) -> dict[str, TemplateBank]:
    banks: dict[str, TemplateBank] = {}
    weights = np.asarray([1.2, 1.2, 1.4, 0.8, 0.35, 0.35, 0.35], dtype=np.float32)
    for label, templates in grouped.items():
        if not templates:
            continue
        features = np.stack([feature_channels(image) for image, _ in templates]).astype(np.float32)
        masks = []
        denominators = []
        for index, _feature in enumerate(features):
            alpha_mask = templates[index][1]
            mask = cv2.morphologyEx(alpha_mask.astype(np.uint8), cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8)).astype(bool)
            if float(mask.mean()) < 0.04:
                mask[:, :] = True
            masks.append(mask)
            denominators.append(max(1.0, float(mask.sum()) * float(weights.sum())))
        banks[label] = TemplateBank(
            label=label,
            features=features,
            masks=np.stack(masks).astype(np.float32),
            denominators=np.asarray(denominators, dtype=np.float32),
            count=len(templates),
        )
    return banks


def patch_to_rgb(patch: Image.Image) -> np.ndarray:
    return np.asarray(patch.convert("RGB").resize(NORMAL_SIZE, Image.Resampling.LANCZOS), dtype=np.uint8)


def score_against(candidate_feature: np.ndarray, bank: TemplateBank) -> float:
    weights = np.asarray([1.2, 1.2, 1.4, 0.8, 0.35, 0.35, 0.35], dtype=np.float32)
    diff = np.abs(bank.features - candidate_feature[None, :, :, :]) * weights
    masked = diff * bank.masks[:, :, :, None]
    scores = 1.0 - (masked.sum(axis=(1, 2, 3)) / bank.denominators)
    if scores.size == 0:
        return 0.0
    return max(0.0, min(1.0, float(scores.max())))


def dark_ratio(candidate: np.ndarray) -> float:
    return collection_visual_stats(candidate)["dark"]


def generic_positive_mask(prototypes: dict[str, TemplateBank]) -> np.ndarray | None:
    cache_key = id(prototypes)
    cached = _GENERIC_POSITIVE_MASK_CACHE.get(cache_key)
    if cached is not None:
        return cached
    masks = [
        bank.masks
        for label, bank in prototypes.items()
        if label in LABELS and bank.masks.size
    ]
    if not masks:
        return None
    stacked = np.concatenate(masks, axis=0)
    mask = np.mean(stacked, axis=0) > 0.18
    if float(mask.mean()) < 0.04:
        mask[:, :] = True
    _GENERIC_POSITIVE_MASK_CACHE[cache_key] = mask
    return mask


def postprocess_direct_label(label: str, scores: dict[str, float], stats: dict[str, float]) -> str:
    adjusted = label
    if (
        adjusted in {"R", "R15"}
        and stats["orange"] >= DIRECT_ORANGE_OVERRIDE
        and stats["orange"] > stats["cyan"] + DIRECT_CYAN_MARGIN
        and stats["cyan"] < DIRECT_ORANGE_CYAN_MAX
    ):
        adjusted = "SSR3" if stats["dark"] >= DIRECT_SSR_DARK_THRESHOLD else "SSR"
    if adjusted in {"SR", "SR15"}:
        is_level_15 = stats["dark"] >= DIRECT_SR_DARK_THRESHOLD
        if (
            is_level_15
            and stats["white"] >= DIRECT_SR_WHITE_GUARD
            and stats["dark_white_ratio"] < DIRECT_SR_DARK_WHITE_RATIO_GUARD
            and (
                stats["purple"] < DIRECT_SR_PURPLE_GUARD
                or stats["active"] < DIRECT_SR_ACTIVE_GUARD
            )
        ):
            is_level_15 = False
        if (
            is_level_15
            and stats["dark"] <= DIRECT_SR_LOW_DARK_GUARD
            and stats["white"] >= DIRECT_SR_BRIGHT_WHITE_GUARD
            and stats["dark_white_ratio"] <= DIRECT_SR_LOW_DARK_WHITE_RATIO_GUARD
            and stats["purple"] <= DIRECT_SR_WEAK_PURPLE_GUARD
            and scores.get("SR15", 0.0) - scores.get("SR", 0.0) <= DIRECT_SR_SCORE_DELTA_GUARD
        ):
            is_level_15 = False
        return "SR15" if is_level_15 else "SR"
    if (
        adjusted in {"SSR", "SSR3"}
        and stats["dark"] >= DIRECT_R15_CYAN_OVERRIDE_DARK_MIN
        and stats["dark"] <= DIRECT_R15_CYAN_OVERRIDE_DARK_MAX
        and stats["white"] <= DIRECT_R15_CYAN_OVERRIDE_WHITE_MAX
        and stats["cyan"] >= stats["orange"] + DIRECT_R15_CYAN_OVERRIDE_MARGIN
        and stats["active"] >= DIRECT_R15_CYAN_OVERRIDE_ACTIVE_MIN
        and scores.get("R15", 0.0) - scores.get("R", 0.0) >= DIRECT_R15_CYAN_OVERRIDE_SCORE_MARGIN
    ):
        return "R15"
    if adjusted in {"SSR", "SSR3"}:
        return "SSR3" if stats["dark"] >= DIRECT_SSR_DARK_THRESHOLD else "SSR"
    if adjusted in {"R", "R15"}:
        return "R15" if stats["dark"] >= DIRECT_R_DARK_THRESHOLD else "R"
    return adjusted


def classify_patch(patch: Image.Image, prototypes: dict[str, TemplateBank], positive_threshold: float, none_margin: float) -> tuple[str, str, str, dict]:
    candidate = patch_to_rgb(patch)
    candidate_feature = feature_channels(candidate)
    color_label = normalize_label(_classify_collection_icon_by_color(patch))
    stats = collection_visual_stats(candidate, generic_positive_mask(prototypes))
    scores = {
        label: score_against(candidate_feature, prototype)
        for label, prototype in prototypes.items()
        if label in LABELS
    }
    none_score = score_against(candidate_feature, prototypes[NONE_LABEL]) if NONE_LABEL in prototypes else 0.0
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    if not ranked:
        return DISPLAY_NONE, DISPLAY_NONE, color_label, {
            "scores": {},
            "best_score": 0.0,
            "second_score": 0.0,
            "none_score": none_score,
            "dark_ratio": stats["dark"],
            "score_margin": 0.0,
            "family_color_ratios": stats,
        }
    best_label, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0
    score_margin = best_score - second_score
    if best_score < positive_threshold or stats["family_max"] < DIRECT_FAMILY_THRESHOLD:
        pure_label = DISPLAY_NONE
    elif none_score and best_score - none_score <= DIRECT_NONE_VETO_MARGIN:
        pure_label = DISPLAY_NONE
    else:
        pure_label = postprocess_direct_label(best_label, scores, stats)
        if (
            pure_label == "R"
            and none_score
            and best_score - none_score <= DIRECT_R_BRIGHT_NONE_MARGIN
            and stats["active"] <= DIRECT_R_BRIGHT_ACTIVE_MAX
            and stats["white"] >= DIRECT_R_BRIGHT_WHITE_MIN
        ):
            pure_label = DISPLAY_NONE
        if (
            pure_label == "R"
            and stats["dark"] <= DIRECT_R_TO_SR_DARK_MAX
            and score_margin <= DIRECT_R_TO_SR_MARGIN
            and color_label == "SR"
        ):
            pure_label = "SR"

    hybrid_label = pure_label

    return pure_label, hybrid_label, color_label, {
        "scores": scores,
        "best_score": best_score,
        "second_score": second_score,
        "none_score": none_score,
        "dark_ratio": stats["dark"],
        "score_margin": score_margin,
        "family_color_ratios": stats,
    }


def slot_patch(row_image: Image.Image, dx: float, dy: float, slot_geometry: SlotGeometry) -> Image.Image:
    y0 = slot_geometry.y_center - slot_geometry.y_half + dy
    y1 = slot_geometry.y_center + slot_geometry.y_half + dy
    return _crop_rel(
        row_image,
        (
            max(0.0, slot_geometry.x_center - slot_geometry.x_half + dx),
            max(0.0, y0),
            min(1.0, slot_geometry.x_center + slot_geometry.x_half + dx),
            min(1.0, y1),
        ),
    )


def classify_slot(
    row_image: Image.Image,
    prototypes: dict[str, TemplateBank],
    threshold: float,
    margin: float,
    slot_geometry: SlotGeometry,
) -> tuple[str, str, str, dict, Image.Image]:
    best: tuple[str, str, str, dict, Image.Image] | None = None
    best_score = -math.inf
    for dx, dy in SEARCH_DELTAS:
        patch = slot_patch(row_image, dx, dy, slot_geometry)
        pure, hybrid, color, debug = classify_patch(patch, prototypes, threshold, margin)
        score = (
            debug["best_score"]
            - max(0.0, debug["none_score"] - debug["best_score"]) * 0.25
            + debug["score_margin"] * 0.08
        )
        if score > best_score:
            best_score = score
            best = (pure, hybrid, color, {**debug, "dx": dx, "dy": dy}, patch)
    assert best is not None
    return best


def slot_geometry_for(
    geometry: CollectionGeometry,
    side: str,
    team_index: int,
    slot_index: int,
    fallback_center: float,
) -> SlotGeometry:
    detailed = (
        geometry.detailed_positions.get(side, {})
        .get(team_index, {})
        .get(slot_index)
    )
    if detailed:
        return detailed
    return SlotGeometry(
        x_center=fallback_center,
        y_center=geometry.y_center,
        x_half=geometry.x_half,
        y_half=geometry.y_half,
    )


def precise_group64_slot_geometry(
    slot_geometry: SlotGeometry,
    side: str,
    match_index: int,
    block_height: int,
) -> tuple[SlotGeometry, float]:
    profile = (
        COLLECTION_PRECISE_GROUP64_WIDE
        if block_height >= COLLECTION_PRECISE_GROUP64_WIDE_BLOCK_HEIGHT_MIN
        else COLLECTION_PRECISE_GROUP64_FHD
    )
    x_half_key = "attacker_x_half" if side == "attacker" else "defender_x_half"
    dy = profile["top_dy"] if match_index in {1, 2} else profile["bottom_dy"]
    return (
        SlotGeometry(
            x_center=slot_geometry.x_center,
            y_center=slot_geometry.y_center,
            x_half=float(profile[x_half_key]),
            y_half=float(profile["y_half"]),
        ),
        float(dy),
    )


def evaluate_dataset(
    dataset: dict,
    boss_labels: dict[tuple[str, int, int], str],
    prototypes: dict[str, TemplateBank],
    threshold: float,
    margin: float,
    geometry: CollectionGeometry,
) -> list[SlotPrediction]:
    players = read_block_players(dataset["excel"])
    predictions: list[SlotPrediction] = []
    with Image.open(dataset["image"]) as image:
        blocks = split_input_image(image, layout="auto", stage_code="group64")
        row_height = (ROW_END - ROW_START) / 5
        for block_index, block in enumerate(blocks[: len(players)]):
            attacker_id, defender_id = players[block_index]
            regions = split_match_block(block.image)
            for side, area, centers, player_id in [
                ("attacker", regions.attacker_area[0], ATTACKER_CARD_SLOT_CENTERS, attacker_id),
                ("defender", regions.defender_area[0], DEFENDER_CARD_SLOT_CENTERS, defender_id),
            ]:
                offsets = geometry.offsets.get(side, geometry.offsets["attacker"])
                for team_index in range(1, 6):
                    row_image = _crop_rel(
                        area,
                        (
                            0.01,
                            ROW_START + (team_index - 1) * row_height,
                            0.99,
                            ROW_START + team_index * row_height,
                        ),
                    )
                    for slot_index in range(1, 6):
                        expected = boss_labels.get((player_id, team_index, slot_index), DISPLAY_NONE)
                        fallback_center = centers[slot_index - 1] - offsets[min(slot_index - 1, len(offsets) - 1)]
                        slot_geometry = slot_geometry_for(geometry, side, team_index, slot_index, fallback_center)
                        slot_geometry, precise_dy = precise_group64_slot_geometry(
                            slot_geometry,
                            side,
                            block.match_index,
                            block.image.height,
                        )
                        slot_geometry = SlotGeometry(
                            x_center=slot_geometry.x_center,
                            y_center=slot_geometry.y_center + precise_dy,
                            x_half=slot_geometry.x_half,
                            y_half=slot_geometry.y_half,
                        )
                        pure, hybrid, color, debug, crop = classify_slot(row_image, prototypes, threshold, margin, slot_geometry)
                        debug = {**debug, "precise_dy": precise_dy}
                        predictions.append(
                            SlotPrediction(
                                dataset=dataset["name"],
                                player_id=player_id,
                                side=side,
                                group_index=block.group_index,
                                match_index=block.match_index,
                                team_index=team_index,
                                slot_index=slot_index,
                                expected=expected,
                                pure_label=pure,
                                hybrid_label=hybrid,
                                color_label=color,
                                best_score=float(debug["best_score"]),
                                second_score=float(debug["second_score"]),
                                none_score=float(debug["none_score"]),
                                dark_ratio=float(debug["dark_ratio"]),
                                crop=crop.copy(),
                                debug=debug,
                            )
                        )
    return predictions


def confusion(predictions: list[SlotPrediction], mode: str) -> Counter:
    matrix: Counter = Counter()
    for item in predictions:
        predicted = prediction_for_mode(item, mode)
        matrix[(item.expected, predicted)] += 1
    return matrix


def prediction_for_mode(item: SlotPrediction, mode: str) -> str:
    if mode == "hybrid":
        return item.hybrid_label
    if mode == "color":
        return item.color_label
    return item.pure_label


def accuracy(predictions: list[SlotPrediction], mode: str) -> float:
    if not predictions:
        return 0.0
    correct = 0
    for item in predictions:
        predicted = prediction_for_mode(item, mode)
        if item.expected == predicted:
            correct += 1
    return correct / len(predictions)


def dataset_accuracy(predictions: list[SlotPrediction]) -> dict:
    datasets = sorted({item.dataset for item in predictions})
    result = {}
    for dataset in datasets:
        items = [item for item in predictions if item.dataset == dataset]
        result[dataset] = {
            "slot_count": len(items),
            "pure_accuracy": accuracy(items, "pure"),
            "hybrid_accuracy": accuracy(items, "hybrid"),
            "color_accuracy": accuracy(items, "color"),
        }
    return result


def write_confusion(path: Path, matrix: Counter) -> None:
    labels = [*LABELS, DISPLAY_NONE]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["expected\\predicted", *labels])
        for expected in labels:
            writer.writerow([expected, *[matrix.get((expected, predicted), 0) for predicted in labels]])


def write_errors(path: Path, predictions: list[SlotPrediction], mode: str) -> list[SlotPrediction]:
    errors = []
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "dataset",
                "player_id",
                "side",
                "group",
                "match",
                "team",
                "slot",
                "expected",
                "predicted",
                "pure",
                "hybrid",
                "color",
                "best_score",
                "second_score",
                "none_score",
                "dark_ratio",
                "dx",
                "dy",
            ]
        )
        for item in predictions:
            predicted = prediction_for_mode(item, mode)
            if item.expected == predicted:
                continue
            errors.append(item)
            writer.writerow(
                [
                    item.dataset,
                    item.player_id,
                    item.side,
                    item.group_index,
                    item.match_index,
                    item.team_index,
                    item.slot_index,
                    item.expected,
                    predicted,
                    item.pure_label,
                    item.hybrid_label,
                    item.color_label,
                    round(item.best_score, 5),
                    round(item.second_score, 5),
                    round(item.none_score, 5),
                    round(item.dark_ratio, 5),
                    item.debug.get("dx"),
                    item.debug.get("dy"),
                ]
            )
    return errors


def make_error_sheet(path: Path, errors: list[SlotPrediction], mode: str, limit: int = 120) -> None:
    cell_w, cell_h = 190, 132
    cols = 5
    rows = max(1, math.ceil(min(len(errors), limit) / cols))
    sheet = Image.new("RGB", (cols * cell_w, rows * cell_h), (246, 246, 246))
    draw = ImageDraw.Draw(sheet)
    for index, item in enumerate(errors[:limit]):
        predicted = prediction_for_mode(item, mode)
        x = (index % cols) * cell_w
        y = (index // cols) * cell_h
        sheet.paste(item.crop.convert("RGB").resize((72, 72), Image.Resampling.NEAREST), (x + 4, y + 4))
        draw.text((x + 82, y + 4), f"{predicted} <- {item.expected}", fill=(0, 0, 0))
        draw.text((x + 82, y + 22), item.dataset, fill=(45, 45, 45))
        draw.text((x + 82, y + 40), f"{item.player_id} {item.side[0]}", fill=(45, 45, 45))
        draw.text((x + 82, y + 58), f"G{item.group_index}M{item.match_index} T{item.team_index}S{item.slot_index}", fill=(45, 45, 45))
        draw.text((x + 4, y + 82), f"score {item.best_score:.3f} none {item.none_score:.3f}", fill=(70, 70, 70))
        draw.text((x + 4, y + 100), f"dark {item.dark_ratio:.3f} dx {item.debug.get('dx')} dy {item.debug.get('dy')}", fill=(70, 70, 70))
    sheet.save(path)


def run(args: argparse.Namespace) -> int:
    root = Path(args.root).resolve() if args.root else project_root()
    template_dir = Path(args.template_dir).resolve() if args.template_dir else tool_root() / "data" / "collection_cv_templates" / "v2_manual"
    output_dir = Path(args.output_dir).resolve() if args.output_dir else template_dir / "evaluation" / datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir.mkdir(parents=True, exist_ok=True)

    boss = read_boss_collections(Path(args.boss_workbook).resolve() if args.boss_workbook else find_boss_workbook(root))
    geometry = load_collection_geometry(template_dir)
    grouped = load_manifest_templates(
        template_dir,
        skip_strong=not args.include_strong_suspects,
        per_label_limit=args.per_label_limit,
    )
    prototypes = make_template_banks(grouped)
    if not all(label in prototypes for label in LABELS):
        missing = [label for label in LABELS if label not in prototypes]
        raise RuntimeError(f"Missing prototypes: {missing}")

    datasets = default_datasets(root)
    all_predictions: list[SlotPrediction] = []
    for dataset in datasets:
        print(f"[dataset] {dataset['name']} image={dataset['image'].name}", flush=True)
        all_predictions.extend(evaluate_dataset(dataset, boss, prototypes, args.positive_threshold, args.none_margin, geometry))

    summary = {
        "template_dir": str(template_dir),
        "prototype_counts": {label: prototype.count for label, prototype in prototypes.items()},
        "geometry": {
            "x_half": geometry.x_half,
            "y_center": geometry.y_center,
            "y_half": geometry.y_half,
            "offsets": {side: list(values) for side, values in geometry.offsets.items()},
            "search_deltas": list(SEARCH_DELTAS),
            "detailed_position_count": sum(
                len(slots)
                for teams in geometry.detailed_positions.values()
                for slots in teams.values()
            ),
        },
        "slot_count": len(all_predictions),
        "pure_accuracy": accuracy(all_predictions, "pure"),
        "hybrid_accuracy": accuracy(all_predictions, "hybrid"),
        "color_accuracy": accuracy(all_predictions, "color"),
        "dataset_accuracy": dataset_accuracy(all_predictions),
        "positive_threshold": args.positive_threshold,
        "family_threshold": DIRECT_FAMILY_THRESHOLD,
        "none_margin": args.none_margin,
        "include_strong_suspects": args.include_strong_suspects,
        "per_label_limit": args.per_label_limit,
    }
    for mode in ("pure", "hybrid", "color"):
        write_confusion(output_dir / f"{mode}_confusion.csv", confusion(all_predictions, mode))
        errors = write_errors(output_dir / f"{mode}_errors.csv", all_predictions, mode)
        make_error_sheet(output_dir / f"{mode}_errors_contact_sheet.png", errors, mode)
        summary[f"{mode}_error_count"] = len(errors)

    (output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    latest = template_dir / "evaluation" / "latest"
    if latest.exists():
        shutil.rmtree(latest)
    shutil.copytree(output_dir, latest)

    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)
    print(f"output={output_dir}", flush=True)
    print(f"latest={latest}", flush=True)
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate NIKKE collection icon OpenCV template matching offline.")
    parser.add_argument("--root", default=None, help="Project root. Defaults to the current arena tool location.")
    parser.add_argument("--template-dir", default=None, help="collection_cv_templates/v2_manual directory.")
    parser.add_argument("--output-dir", default=None, help="Evaluation output directory.")
    parser.add_argument("--boss-workbook", default=None, help="Reference workbook with collection labels.")
    parser.add_argument("--positive-threshold", type=float, default=DIRECT_SCORE_THRESHOLD)
    parser.add_argument("--none-margin", type=float, default=0.03)
    parser.add_argument("--per-label-limit", type=int, default=64)
    parser.add_argument("--include-strong-suspects", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
