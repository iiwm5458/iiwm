from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import cv2
import numpy as np
from openpyxl import load_workbook
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from recognizer.arena_ocr import ArenaOCRRecognizer, OCRItem
from recognizer.image_preprocess import prepare_for_ocr
from recognizer.image_splitter import split_input_image, split_match_block
from recognizer.result_parser import (
    ATTACKER_POWER_SLOT_CENTERS,
    DEFENDER_POWER_SLOT_CENTERS,
    MAX_CARD_POWER,
    MIN_CARD_POWER,
    _crop_rel,
    _extract_card_power,
)


ROW_START = 0.275
ROW_END = 0.925
POWER_CROP_TOP = 0.64
POWER_CROP_BOTTOM = 0.985
POWER_CROP_LEFT = 0.125
POWER_CROP_RIGHT = 0.125
POWER_TEMPLATE_SIZE = (18, 28)


@dataclass(frozen=True)
class ExpectedSlot:
    player_id: str
    nickname: str
    group_index: int
    seat_index: int
    team_index: int
    slot_index: int
    name: str
    power: int


@dataclass
class DigitTemplate:
    sample_id: str
    digit: str
    mask: np.ndarray


@dataclass
class CompiledDigitTemplates:
    sample_ids: list[str]
    masks: np.ndarray
    sums: np.ndarray


@dataclass
class DigitPrototype:
    digit: str
    vector: np.ndarray


@dataclass
class PowerSample:
    sample_id: str
    image_name: str
    group_index: int
    match_index: int
    side: str
    player_id: str
    nickname: str
    team_index: int
    slot_index: int
    name: str
    expected: int
    raw_crop: Image.Image
    digit_band: Image.Image
    masks: list[np.ndarray]
    segment_count: int


@dataclass(frozen=True)
class PaddleStripReading:
    text: str = ""
    confidence: float = 0.0
    value: int | None = None


def normalize_player_id(value) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D+", "", text)
    return digits.zfill(8) if digits else ""


def to_int(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    match = re.search(r"\d+", str(value).replace(",", ""))
    return int(match.group(0)) if match else None


def parse_group_seat(value) -> tuple[int, int] | None:
    text = str(value or "")
    match = re.search(r"\u7b2c\s*(\d+)\s*\u7ec4\s*(\d+)\s*\u53f7\u4f4d", text)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def read_expected_slots(path: Path) -> dict[tuple[int, int, int, int], ExpectedSlot]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    team_counter: dict[str, int] = defaultdict(int)
    slots: dict[tuple[int, int, int, int], ExpectedSlot] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        player_id = normalize_player_id(row[0])
        seat = parse_group_seat(row[2])
        if not player_id or not seat:
            continue
        group_index, seat_index = seat
        team_counter[player_id] += 1
        team_index = team_counter[player_id]
        nickname = str(row[1] or "")
        for slot_index in range(1, 6):
            name = str(row[14 + (slot_index - 1) * 2] or "")
            power = to_int(row[15 + (slot_index - 1) * 2])
            if power is None:
                continue
            slots[(group_index, seat_index, team_index, slot_index)] = ExpectedSlot(
                player_id=player_id,
                nickname=nickname,
                group_index=group_index,
                seat_index=seat_index,
                team_index=team_index,
                slot_index=slot_index,
                name=name,
                power=power,
            )
    workbook.close()
    return slots


def seat_for_block(match_index: int, side: str) -> int:
    base = (match_index - 1) * 2 + 1
    return base if side == "attacker" else base + 1


def normalize_digit_mask(mask: np.ndarray) -> np.ndarray:
    target_w, target_h = POWER_TEMPLATE_SIZE
    if mask.size == 0 or not np.any(mask):
        return np.zeros((target_h, target_w), dtype=np.uint8)
    ys, xs = np.where(mask > 0)
    trimmed = mask[int(ys.min()) : int(ys.max()) + 1, int(xs.min()) : int(xs.max()) + 1]
    h, w = trimmed.shape[:2]
    scale = min((target_w - 4) / max(1, w), (target_h - 4) / max(1, h))
    scaled_w = max(1, int(round(w * scale)))
    scaled_h = max(1, int(round(h * scale)))
    resized = Image.fromarray((trimmed * 255).astype(np.uint8)).resize((scaled_w, scaled_h), Image.Resampling.NEAREST)
    canvas = Image.new("L", (target_w, target_h), 0)
    canvas.paste(resized, ((target_w - scaled_w) // 2, (target_h - scaled_h) // 2))
    return (np.asarray(canvas) > 0).astype(np.uint8)


def split_wide_component(mask: np.ndarray, split_threshold: int) -> list[np.ndarray]:
    h, w = mask.shape[:2]
    if w <= split_threshold or w < 14 or h < 8:
        return [mask]
    projection = mask.sum(axis=0)
    start = max(3, int(round(w * 0.35)))
    end = min(w - 3, int(round(w * 0.65)))
    if start >= end:
        return [mask]
    cut = min(range(start, end + 1), key=lambda x: (int(projection[x]), abs(x - w / 2.0)))
    left = mask[:, :cut]
    right = mask[:, cut:]
    if not np.any(left) or not np.any(right):
        return [mask]
    return [left, right]


def segment_digits(image: Image.Image, expected_digits: int | None = None) -> list[np.ndarray]:
    gray = np.asarray(image.convert("L"), dtype=np.uint8)
    if gray.size == 0:
        return []
    # Digits are usually the largest dark connected components in the lower
    # half of the crop. The left sword icon is kept out by taking rightmost
    # expected components when the expected length is known.
    threshold = max(135, min(220, int(np.percentile(gray, 46))))
    mask = (gray < threshold).astype(np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, np.ones((2, 2), np.uint8))
    component_count, labels, stats, _centroids = cv2.connectedComponentsWithStats(mask, 8)
    height, width = gray.shape[:2]
    min_area = max(5, int(round(width * height * 0.0015)))
    min_height = max(7, int(round(height * 0.16)))
    components: list[tuple[int, np.ndarray]] = []
    for component_index in range(1, component_count):
        x, y, w, h, area = stats[component_index]
        if area < min_area or h < min_height or w < 3:
            continue
        if y + h < height * 0.35:
            continue
        component = (labels[y : y + h, x : x + w] == component_index).astype(np.uint8)
        components.append((int(x), component))
    if not components:
        return []
    widths = sorted(comp.shape[1] for _x, comp in components if comp.shape[1] >= 3)
    split_threshold = max(18, int(round((np.median(widths) if widths else 9) * 1.85)))
    expanded: list[tuple[int, np.ndarray]] = []
    for x, component in sorted(components, key=lambda item: item[0]):
        parts = split_wide_component(component, split_threshold)
        part_x = x
        for part in parts:
            if np.any(part):
                expanded.append((part_x, normalize_digit_mask(part)))
            part_x += max(1, part.shape[1])
    expanded = [(x, mask) for x, mask in expanded if np.any(mask)]
    expanded.sort(key=lambda item: item[0])
    if expected_digits is not None and len(expanded) >= expected_digits:
        expanded = expanded[-expected_digits:]
    return [mask for _x, mask in expanded]


def isolate_power_band(image: Image.Image) -> Image.Image:
    gray = np.asarray(image.convert("L"), dtype=np.uint8)
    if gray.size == 0:
        return image
    height, width = gray.shape[:2]
    mask = ((gray < 185) & (gray > 35)).astype(np.uint8)
    mask[: int(height * 0.35), :] = 0
    mask[int(height * 0.80) :, :] = 0
    row_projection = mask.sum(axis=1)
    active_rows = np.where(row_projection > max(4, int(round(width * 0.025))))[0]
    if active_rows.size == 0:
        return image.crop((0, int(height * 0.38), width, int(height * 0.78)))

    groups: list[tuple[int, int]] = []
    start = previous = int(active_rows[0])
    for row in active_rows[1:]:
        row = int(row)
        if row <= previous + 1:
            previous = row
            continue
        groups.append((start, previous))
        start = previous = row
    groups.append((start, previous))

    best_start, best_end = max(
        groups,
        key=lambda item: int(row_projection[item[0] : item[1] + 1].sum()),
    )
    y0 = max(0, best_start - 8)
    y1 = min(height, best_end + 9)
    band_mask = ((gray[y0:y1] < 185) & (gray[y0:y1] > 35)).astype(np.uint8)
    col_projection = band_mask.sum(axis=0)
    active_cols = np.where(col_projection > 0)[0]
    if active_cols.size:
        x0 = max(0, int(active_cols[0]) - 8)
        x1 = min(width, int(active_cols[-1]) + 9)
    else:
        x0, x1 = 0, width
    return image.crop((x0, y0, x1, y1))


def _best_paddle_strip_reading(items: list[OCRItem]) -> PaddleStripReading:
    candidates: list[PaddleStripReading] = []
    for item in items:
        value = _extract_card_power(item.text)
        candidates.append(PaddleStripReading(text=item.text, confidence=item.confidence, value=value))
    if not candidates:
        return PaddleStripReading()
    return max(
        candidates,
        key=lambda item: (
            item.value is not None,
            item.confidence,
            len(str(item.value)) if item.value is not None else 0,
        ),
    )


def recognize_paddle_strip(
    sample: PowerSample,
    ocr: ArenaOCRRecognizer,
) -> tuple[PaddleStripReading, PaddleStripReading, PaddleStripReading, int | None, int | None, float]:
    start = time.perf_counter()
    native = _best_paddle_strip_reading(
        ocr.recognize_text_line(sample.digit_band, f"{sample.sample_id}_power_strip_native")
    )
    prepared = _best_paddle_strip_reading(
        ocr.recognize_text_line(prepare_for_ocr(sample.digit_band), f"{sample.sample_id}_power_strip_prepared")
    )
    detected = _best_paddle_strip_reading(
        ocr.recognize_region(prepare_for_ocr(sample.digit_band), f"{sample.sample_id}_power_strip_detect")
    )
    line_consensus = native.value if native.value is not None and native.value == prepared.value else None
    consensus = line_consensus
    if detected.value is not None and line_consensus is not None and len(str(detected.value)) > len(str(line_consensus)):
        consensus = None
    return native, prepared, detected, line_consensus, consensus, time.perf_counter() - start


def digit_similarity(left: np.ndarray, right: np.ndarray) -> float:
    intersection = float(np.logical_and(left, right).sum())
    denominator = float(left.sum() + right.sum())
    return 2.0 * intersection / denominator if denominator else 0.0


def extract_power_crop(row_image: Image.Image, center: float) -> Image.Image:
    return _crop_rel(
        row_image,
        (
            max(0.0, center - POWER_CROP_LEFT),
            POWER_CROP_TOP,
            min(1.0, center + POWER_CROP_RIGHT),
            POWER_CROP_BOTTOM,
        ),
    )


def collect_samples(image_path: Path, expected: dict[tuple[int, int, int, int], ExpectedSlot]) -> list[PowerSample]:
    image = Image.open(image_path).convert("RGB")
    blocks = split_input_image(image, layout="auto", stage_code="group64")
    samples: list[PowerSample] = []
    for block in blocks:
        regions = split_match_block(block.image)
        for side, side_image, centers in (
            ("attacker", regions.attacker_area[0], ATTACKER_POWER_SLOT_CENTERS),
            ("defender", regions.defender_area[0], DEFENDER_POWER_SLOT_CENTERS),
        ):
            seat_index = seat_for_block(block.match_index, side)
            row_h = (ROW_END - ROW_START) / 5
            for team_index in range(1, 6):
                y0 = ROW_START + (team_index - 1) * row_h
                y1 = ROW_START + team_index * row_h
                row_image = _crop_rel(side_image, (0.01, y0, 0.99, y1))
                for slot_index, center in enumerate(centers, start=1):
                    slot = expected.get((block.group_index, seat_index, team_index, slot_index))
                    if not slot:
                        continue
                    raw_crop = extract_power_crop(row_image, center)
                    digit_band = isolate_power_band(raw_crop)
                    expected_digits = len(str(slot.power))
                    masks = segment_digits(digit_band, expected_digits=expected_digits)
                    samples.append(
                        PowerSample(
                            sample_id=(
                                f"{image_path.stem}:G{block.group_index}:M{block.match_index}:"
                                f"{side}:T{team_index}:S{slot_index}"
                            ),
                            image_name=image_path.name,
                            group_index=block.group_index,
                            match_index=block.match_index,
                            side=side,
                            player_id=slot.player_id,
                            nickname=slot.nickname,
                            team_index=team_index,
                            slot_index=slot_index,
                            name=slot.name,
                            expected=slot.power,
                            raw_crop=raw_crop,
                            digit_band=digit_band,
                            masks=masks,
                            segment_count=len(masks),
                        )
                    )
    return samples


def build_templates(samples: list[PowerSample]) -> dict[str, list[DigitTemplate]]:
    templates: dict[str, list[DigitTemplate]] = defaultdict(list)
    for sample in samples:
        digits = str(sample.expected)
        if len(sample.masks) != len(digits):
            continue
        for digit, mask in zip(digits, sample.masks):
            templates[digit].append(DigitTemplate(sample_id=sample.sample_id, digit=digit, mask=mask))
    return templates


def compile_templates(templates: dict[str, list[DigitTemplate]]) -> dict[str, CompiledDigitTemplates]:
    compiled: dict[str, CompiledDigitTemplates] = {}
    for digit, items in templates.items():
        if not items:
            continue
        masks = np.stack([item.mask.astype(np.uint8) for item in items], axis=0)
        compiled[digit] = CompiledDigitTemplates(
            sample_ids=[item.sample_id for item in items],
            masks=masks,
            sums=masks.reshape(masks.shape[0], -1).sum(axis=1).astype(np.float32),
        )
    return compiled


def compile_prototypes(templates: dict[str, list[DigitTemplate]]) -> dict[str, DigitPrototype]:
    prototypes: dict[str, DigitPrototype] = {}
    for digit, items in templates.items():
        if not items:
            continue
        vectors = np.stack([item.mask.astype(np.float32).reshape(-1) for item in items], axis=0)
        prototypes[digit] = DigitPrototype(digit=digit, vector=vectors.mean(axis=0))
    return prototypes


def best_digit_score(mask: np.ndarray, bank: CompiledDigitTemplates, sample_id: str) -> float:
    if bank.masks.size == 0:
        return 0.0
    mask_u8 = mask.astype(np.uint8)
    mask_sum = float(mask_u8.sum())
    if mask_sum <= 0:
        return 0.0
    intersections = np.logical_and(bank.masks, mask_u8).reshape(bank.masks.shape[0], -1).sum(axis=1).astype(np.float32)
    denominators = bank.sums + mask_sum
    scores = np.divide(intersections * 2.0, denominators, out=np.zeros_like(intersections), where=denominators > 0)
    same_sample = np.asarray([sid == sample_id for sid in bank.sample_ids], dtype=bool)
    if np.any(~same_sample):
        scores[same_sample] = -1.0
    return float(scores.max()) if scores.size else 0.0


def recognize_sample(sample: PowerSample, templates: dict[str, CompiledDigitTemplates]) -> tuple[int | None, dict]:
    if len(sample.masks) not in {5, 6}:
        return None, {"reason": "bad_segment_count", "segment_count": len(sample.masks)}
    digits: list[str] = []
    digit_scores: list[float] = []
    margins: list[float] = []
    per_digit: list[dict] = []
    for mask in sample.masks:
        scores: list[tuple[float, str]] = []
        for digit, digit_templates in templates.items():
            scores.append((best_digit_score(mask, digit_templates, sample.sample_id), digit))
        scores.sort(reverse=True)
        if not scores:
            return None, {"reason": "no_templates"}
        best_score, best_digit = scores[0]
        second_score = scores[1][0] if len(scores) > 1 else 0.0
        digits.append(best_digit)
        digit_scores.append(float(best_score))
        margins.append(float(best_score - second_score))
        per_digit.append({"digit": best_digit, "score": best_score, "second": second_score})
    if digits[0] == "0":
        return None, {"reason": "leading_zero", "digits": digits}
    value = int("".join(digits))
    if not MIN_CARD_POWER <= value <= MAX_CARD_POWER:
        return None, {"reason": "range", "value": value}
    avg_score = sum(digit_scores) / len(digit_scores)
    min_score = min(digit_scores)
    min_margin = min(margins)
    accepted = bool(min_score >= 0.70 and avg_score >= 0.84 and min_margin >= -0.02)
    return value, {
        "accepted": accepted,
        "avg_score": avg_score,
        "min_score": min_score,
        "min_margin": min_margin,
        "per_digit": per_digit,
        "segment_count": len(sample.masks),
    }


def recognize_sample_prototype(
    sample: PowerSample,
    prototypes: dict[str, DigitPrototype],
    accept_margin: float,
) -> tuple[int | None, dict]:
    if len(sample.masks) not in {5, 6}:
        return None, {"reason": "bad_segment_count", "segment_count": len(sample.masks)}
    digits: list[str] = []
    distances: list[float] = []
    margins: list[float] = []
    per_digit: list[dict] = []
    for mask in sample.masks:
        vector = mask.astype(np.float32).reshape(-1)
        scores: list[tuple[float, str]] = []
        for digit, prototype in prototypes.items():
            distance = float(np.mean((vector - prototype.vector) ** 2))
            scores.append((distance, digit))
        scores.sort()
        if not scores:
            return None, {"reason": "no_templates"}
        best_distance, best_digit = scores[0]
        second_distance = scores[1][0] if len(scores) > 1 else best_distance
        digits.append(best_digit)
        distances.append(best_distance)
        margins.append(float(second_distance - best_distance))
        per_digit.append({"digit": best_digit, "distance": best_distance, "second": second_distance})
    if digits[0] == "0":
        return None, {"reason": "leading_zero", "digits": digits}
    value = int("".join(digits))
    if not MIN_CARD_POWER <= value <= MAX_CARD_POWER:
        return None, {"reason": "range", "value": value}
    min_margin = min(margins)
    avg_distance = sum(distances) / len(distances)
    max_distance = max(distances)
    return value, {
        "accepted": min_margin >= accept_margin,
        "avg_score": 1.0 - avg_distance,
        "min_score": 1.0 - max_distance,
        "min_margin": min_margin,
        "per_digit": per_digit,
        "segment_count": len(sample.masks),
    }


def write_outputs(
    samples: list[PowerSample],
    templates: dict[str, list[DigitTemplate]],
    output_dir: Path,
    save_crops: bool,
    classifier: str,
    accept_margin: float,
    paddle_strip_ocr: ArenaOCRRecognizer | None = None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    compiled_templates = compile_templates(templates)
    prototypes = compile_prototypes(templates)
    rows: list[dict] = []
    errors: list[dict] = []
    paddle_elapsed_seconds = 0.0
    crop_dir = output_dir / "power_crops"
    if save_crops:
        crop_dir.mkdir(exist_ok=True)
    for sample in samples:
        if classifier == "prototype":
            value, debug = recognize_sample_prototype(sample, prototypes, accept_margin=accept_margin)
        else:
            value, debug = recognize_sample(sample, compiled_templates)
        accepted = bool(debug.get("accepted")) if value is not None else False
        correct = value == sample.expected
        native_reading = PaddleStripReading()
        prepared_reading = PaddleStripReading()
        detected_reading = PaddleStripReading()
        paddle_line_consensus: int | None = None
        paddle_consensus: int | None = None
        if paddle_strip_ocr is not None:
            (
                native_reading,
                prepared_reading,
                detected_reading,
                paddle_line_consensus,
                paddle_consensus,
                elapsed,
            ) = recognize_paddle_strip(sample, paddle_strip_ocr)
            paddle_elapsed_seconds += elapsed
        row = {
            "image": sample.image_name,
            "sample_id": sample.sample_id,
            "group": sample.group_index,
            "match": sample.match_index,
            "side": sample.side,
            "player_id": sample.player_id,
            "nickname": sample.nickname,
            "team": sample.team_index,
            "slot": sample.slot_index,
            "name": sample.name,
            "expected": sample.expected,
            "cv_value": value,
            "accepted": accepted,
            "correct": correct,
            "segment_count": sample.segment_count,
            "avg_score": debug.get("avg_score"),
            "min_score": debug.get("min_score"),
            "min_margin": debug.get("min_margin"),
            "reason": debug.get("reason", ""),
            "paddle_native_text": native_reading.text,
            "paddle_native_confidence": native_reading.confidence,
            "paddle_native_value": native_reading.value,
            "paddle_prepared_text": prepared_reading.text,
            "paddle_prepared_confidence": prepared_reading.confidence,
            "paddle_prepared_value": prepared_reading.value,
            "paddle_detected_text": detected_reading.text,
            "paddle_detected_confidence": detected_reading.confidence,
            "paddle_detected_value": detected_reading.value,
            "paddle_line_consensus_value": paddle_line_consensus,
            "paddle_consensus_value": paddle_consensus,
            "paddle_consensus_correct": paddle_consensus == sample.expected if paddle_consensus is not None else False,
        }
        rows.append(row)
        if value != sample.expected:
            errors.append(row)
            if save_crops:
                safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", sample.sample_id)
                stem = f"{safe_name}_expected_{sample.expected}_cv_{value}"
                sample.raw_crop.save(crop_dir / f"{stem}_raw.png")
                sample.digit_band.save(crop_dir / f"{stem}_band.png")

    fieldnames = list(rows[0].keys()) if rows else []
    for filename, data in (("power_cv_eval_slots.csv", rows), ("power_cv_eval_errors.csv", errors)):
        with (output_dir / filename).open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

    accepted_rows = [row for row in rows if row["accepted"]]
    raw_correct = sum(1 for row in rows if row["correct"])
    accepted_correct = sum(1 for row in accepted_rows if row["correct"])
    paddle_native_rows = [row for row in rows if row["paddle_native_value"] is not None]
    paddle_prepared_rows = [row for row in rows if row["paddle_prepared_value"] is not None]
    paddle_detected_rows = [row for row in rows if row["paddle_detected_value"] is not None]
    paddle_line_consensus_rows = [row for row in rows if row["paddle_line_consensus_value"] is not None]
    paddle_consensus_rows = [row for row in rows if row["paddle_consensus_value"] is not None]
    paddle_native_correct = sum(row["paddle_native_value"] == row["expected"] for row in paddle_native_rows)
    paddle_prepared_correct = sum(row["paddle_prepared_value"] == row["expected"] for row in paddle_prepared_rows)
    paddle_detected_correct = sum(row["paddle_detected_value"] == row["expected"] for row in paddle_detected_rows)
    paddle_line_consensus_correct = sum(
        row["paddle_line_consensus_value"] == row["expected"] for row in paddle_line_consensus_rows
    )
    paddle_consensus_correct = sum(row["paddle_consensus_correct"] for row in paddle_consensus_rows)
    summary = {
        "classifier": classifier,
        "accept_margin": accept_margin if classifier == "prototype" else None,
        "total_slots": len(rows),
        "segmented_slots": sum(1 for row in rows if row["segment_count"] in {5, 6}),
        "template_counts": {digit: len(items) for digit, items in sorted(templates.items())},
        "raw_correct": raw_correct,
        "raw_errors": len(errors),
        "raw_accuracy": raw_correct / len(rows) if rows else None,
        "accepted": len(accepted_rows),
        "accepted_correct": accepted_correct,
        "accepted_errors": len(accepted_rows) - accepted_correct,
        "accepted_accuracy": accepted_correct / len(accepted_rows) if accepted_rows else None,
        "rejected": len(rows) - len(accepted_rows),
        "segment_count_distribution": Counter(str(row["segment_count"]) for row in rows),
        "paddle_strip_enabled": paddle_strip_ocr is not None,
        "paddle_strip_engine": paddle_strip_ocr.engine_name if paddle_strip_ocr is not None else "",
        "paddle_strip_seconds": paddle_elapsed_seconds,
        "paddle_strip_ms_per_slot": paddle_elapsed_seconds * 1000.0 / len(rows) if rows else None,
        "paddle_native_valid": len(paddle_native_rows),
        "paddle_native_correct": paddle_native_correct,
        "paddle_native_accuracy": paddle_native_correct / len(paddle_native_rows) if paddle_native_rows else None,
        "paddle_prepared_valid": len(paddle_prepared_rows),
        "paddle_prepared_correct": paddle_prepared_correct,
        "paddle_prepared_accuracy": paddle_prepared_correct / len(paddle_prepared_rows) if paddle_prepared_rows else None,
        "paddle_detected_valid": len(paddle_detected_rows),
        "paddle_detected_correct": paddle_detected_correct,
        "paddle_detected_accuracy": paddle_detected_correct / len(paddle_detected_rows) if paddle_detected_rows else None,
        "paddle_line_consensus": len(paddle_line_consensus_rows),
        "paddle_line_consensus_correct": paddle_line_consensus_correct,
        "paddle_line_consensus_accuracy": (
            paddle_line_consensus_correct / len(paddle_line_consensus_rows) if paddle_line_consensus_rows else None
        ),
        "paddle_consensus": len(paddle_consensus_rows),
        "paddle_consensus_correct": paddle_consensus_correct,
        "paddle_consensus_accuracy": paddle_consensus_correct / len(paddle_consensus_rows) if paddle_consensus_rows else None,
        "error_examples": errors[:50],
    }
    (output_dir / "power_cv_eval_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Offline evaluator for card power CV recognition.")
    parser.add_argument("--baseline", required=True, type=Path, help="Boss baseline workbook.")
    parser.add_argument("--image", action="append", required=True, type=Path, help="64进32 detailed composite image. Repeatable.")
    parser.add_argument("--output-dir", required=True, type=Path, help="Directory for CSV/JSON reports.")
    parser.add_argument("--save-crops", action="store_true", help="Save crops for mismatched CV predictions.")
    parser.add_argument("--classifier", choices=("prototype", "dice"), default="prototype")
    parser.add_argument("--accept-margin", type=float, default=0.012)
    parser.add_argument("--paddle-strip", action="store_true", help="Evaluate Paddle recognition-only on each isolated power strip.")
    parser.add_argument("--gpu", action="store_true", help="Use the configured Paddle GPU runtime for --paddle-strip.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    expected = read_expected_slots(args.baseline)
    all_samples: list[PowerSample] = []
    for image_path in args.image:
        all_samples.extend(collect_samples(image_path, expected))
    templates = build_templates(all_samples)
    paddle_strip_ocr = None
    if args.paddle_strip:
        paddle_strip_ocr = ArenaOCRRecognizer(use_gpu=args.gpu)
        if not paddle_strip_ocr.available:
            print(paddle_strip_ocr.error or "PaddleOCR is unavailable for strip evaluation.", file=sys.stderr)
            return 2
    write_outputs(
        all_samples,
        templates,
        args.output_dir,
        save_crops=args.save_crops,
        classifier=args.classifier,
        accept_margin=args.accept_margin,
        paddle_strip_ocr=paddle_strip_ocr,
    )
    print(f"samples={len(all_samples)}")
    if paddle_strip_ocr is not None:
        print(f"paddle_strip_engine={paddle_strip_ocr.engine_name}")
    print(f"output={args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
